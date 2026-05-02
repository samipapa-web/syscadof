# script_analyse_type.py

import os
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
from psycopg2.extras import RealDictCursor

# ==============================
# VARIABLES INJECTÉES
# ==============================
source_a_id = globals().get("source_a_id")
source_b_id = globals().get("source_b_id")
SORTIE_DIR = globals().get("SORTIE_DIR")
CLEAN_DIR = globals().get("CLEAN_DIR")
session = globals().get("session", {})
db_conn = globals().get("db_conn")

if not db_conn:
    raise Exception("db_conn non injecté")

if not CLEAN_DIR:
    raise Exception("CLEAN_DIR non défini")

if not SORTIE_DIR:
    raise Exception("SORTIE_DIR non défini")

if not source_a_id or not source_b_id:
    raise Exception("IDs sources obligatoires")

utilisateur = session.get("user", "inconnu")


# ==============================
# DB HELPERS (ROBUSTES)
# ==============================
def get_source_info(source_id):
    with db_conn() as conn:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:

            cur.execute("""
                SELECT nom_fichier, annee, mois
                FROM sources_clean
                WHERE id = %s
            """, (source_id,))

            row = cur.fetchone()

            if not row:
                raise Exception(f"Source introuvable : {source_id}")

            return row


# ==============================
# FICHIERS
# ==============================
info_a = get_source_info(source_a_id)
info_b = get_source_info(source_b_id)

file_a = os.path.join(CLEAN_DIR, info_a["nom_fichier"])
file_b = os.path.join(CLEAN_DIR, info_b["nom_fichier"])

if not os.path.exists(file_a):
    raise FileNotFoundError(file_a)

if not os.path.exists(file_b):
    raise FileNotFoundError(file_b)

os.makedirs(SORTIE_DIR, exist_ok=True)


# ==============================
# LECTURE
# ==============================
df_a = pd.read_excel(file_a)
df_b = pd.read_excel(file_b)

df_a.columns = df_a.columns.str.strip()
df_b.columns = df_b.columns.str.strip()


# ====================================
# ALGORYTHME DE CROISEMENT MODIFIABLE
# ====================================

if "NIU" not in df_a.columns or "NIU" not in df_b.columns:
    raise Exception("Colonne NIU obligatoire")

# JOINTURE
# ==============================
df = df_a.merge(df_b, on="NIU", how="left")

# TRAITEMENT
# ==============================
df["Total général"] = pd.to_numeric(df["Total général"], errors="coerce")
df["achats_hors_region_march"] = pd.to_numeric(df["achats_hors_region_march"], errors="coerce")

df = df[df["achats_hors_region_march"].notna()]
df = df[df["achats_hors_region_march"] != 0]

df["ecart"] = df["Total général"] - df["achats_hors_region_march"]

df = df[df["ecart"] >= 10_000_000]
df.sort_values(by="ecart", ascending=False, inplace=True)

suffix_output = "ecart_mininal_10M"


# ====================================
# FIN DE L'ALGORYTHME DE CROISEMENT MODIFIABLE
# ====================================

# ==============================
# SORTIE
# ==============================
nom_fichier = f"matching_{source_a_id}X{source_b_id}_{suffix_output}.xlsx"
chemin = os.path.join(SORTIE_DIR, nom_fichier)
df.to_excel(chemin, index=False)

# ==============================
# FORMAT
# ==============================
wb = load_workbook(chemin)
ws = wb.active

for col in ["Total général", "achats_hors_region_march", "ecart"]:
    if col in df.columns:
        idx = df.columns.get_loc(col) + 1
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=idx).number_format = '# ##0'

wb.save(chemin)


# ==============================
# RESULTAT OBLIGATOIRE
# ==============================
dernier_croisement = {
    "id_source_a": int(source_a_id),
    "id_source_b": int(source_b_id),
    "nom_fichier": nom_fichier,
    "type_fichier": "xlsx",
    "utilisateur": utilisateur,
    "date_stockage": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
}
